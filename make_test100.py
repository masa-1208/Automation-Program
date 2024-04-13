import openpyxl as excel

book = excel.Workbook()
sheet = book.active

for row in range(1, 101):
    for col in range(1,101):
        cell = sheet.cell(row, col)
        cell.value = excel.utils.get_column_letter(col) + str(row)

book.save("test100.xlsx")