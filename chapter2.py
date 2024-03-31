import openpyxl as excel

wareki_table = [
    {"name": "Meiji", "start": 1868, "end": 1912},
    {"name": "Taisho", "start": 1912, "end": 1926},
    {"name": "Showa", "start": 1926, "end": 1989},
    {"name": "Heisei", "start": 1989, "end": 2019},
    {"name": "Reiwa", "start": 2019, "end": 9999}
]

# Function to convert from Western to Japanese calendar
def seireki_wareki(year):
    for w in wareki_table:
        if w["start"] <= year < w["end"]:
            y = "year" + str(year - w["start"] + 1)
            return w["name"] + y
    return "unknown"

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "Western"
sheet["B1"] = "Japanese Calendar"

start_y = 1930
for i in range(101):
    sei = start_y + i
    wa = seireki_wareki(sei)

    sheet.cell(i+2, 1, value="year:" + str(sei))
    sheet.cell(i+2, 2, value=wa)

book.save("wareki.xlsx")