import openpyxl as excel

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "西暦"
sheet["B1"] = "和暦"

start_year = 1930
for i in range(100):
    sei = str(start_year + i)
    wa = '=TEXT("{}/1/1", "ggge年")'.format(sei)

    sheet.cell(row=(2+i), column=1, value=sei+'年')
    sheet.cell(row=(2+i), column=2, value=wa)
    print(sei, "=", wa)

book.save("warekist2.xlsx")